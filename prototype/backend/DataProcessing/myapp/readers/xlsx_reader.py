import pandas as pd
import openpyxl

def read_xlsx(file_path):
    """专门处理 .xlsx 文件，保留公式错误。"""
    sheets_dfs = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    for ws in wb.worksheets:
        if ws.max_row < 4: continue
        header = [cell.value for cell in ws[4]]
        data_rows = []
        for row in ws.iter_rows(min_row=5):
            row_data = [str(cell.value) if cell.data_type == 'e' else cell.value for cell in row]
            data_rows.append(row_data)
        
        if data_rows:
            sheets_dfs.append(pd.DataFrame(data_rows, columns=header))
    return sheets_dfs